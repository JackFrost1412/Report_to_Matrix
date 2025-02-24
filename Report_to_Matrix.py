import pandas as pd
from openpyxl import load_workbook

# Hàm tìm version tiếp theo của sheet hiện có (ví dụ: List DIM_v1, v2,...)
def get_next_version(sheet_name, existing_sheets):
    version = 0
    base_name = sheet_name
    while f"{base_name}_v{version+1}" in existing_sheets:
        version += 1
    return f"{base_name}_v{version+1}"

# Hàm chuẩn hóa toàn bộ DataFrame: chuẩn hóa các chuỗi thành dạng title case
def normalize_dataframe(df):
    columns_to_normalize = ['Tên Chiều/Chỉ tiêu', 'Chiều cơ sở', 'Chỉ tiêu cơ sở', 'Chi tiết']
    
    for column in columns_to_normalize:
        df[column] = df[column].apply(lambda value: value.title() if isinstance(value, str) else value)
    return df

def report_to_matrix(file_path):
    sheet_dict = pd.read_excel(file_path, sheet_name=None)

    data_dim = []
    data_mea = []
    list_dim = pd.DataFrame(columns=['DIM name', 'Thuộc tính'])
    list_mea = pd.DataFrame(columns=['MEA name', 'Phân loại', 'MEA cơ sở'])

    rpt_dim_dict = {}
    rpt_mea_dict = {}
    mea_dim_dict = {}

    for sheet_name, df in sheet_dict.items():
        if sheet_name.startswith("BA"):
            # Chuẩn hóa lại DataFrame
            df = normalize_dataframe(df)

            for rpt_id in df["Báo cáo"].unique():
                rpt_dim_dict[rpt_id] = []
                rpt_mea_dict[rpt_id] = []
                mea_dim_dict[rpt_id] = []

            for index, row in df.iterrows():
                if row["Loại"] == "Chiều" or row["Loại"] == "Thuộc tính":
                    data_dim.append({
                        'DIM name': row['Chiều cơ sở'],
                        'Thuộc tính': row['Tên Chiều/Chỉ tiêu']
                    })
                    rpt_dim_dict[rpt_id].append(row["Chiều cơ sở"])

                elif row["Loại"] == "Chỉ tiêu":
                    data_mea.append({
                        'MEA name': row['Chỉ tiêu cơ sở'],
                        'Phân loại': 'Chỉ tiêu cơ sở',
                    })
                    rpt_mea_dict[rpt_id].append(row["Chỉ tiêu cơ sở"])

                elif row["Loại"] == "Chỉ tiêu phái sinh" and pd.notna(row["Chỉ tiêu cơ sở"]):
                    data_mea.append({
                        'MEA name': row['Tên Chiều/Chỉ tiêu'],
                        'Phân loại': 'Chỉ tiêu phái sinh',
                        'MEA cơ sở': row['Chỉ tiêu cơ sở']
                    })
                    rpt_mea_dict[rpt_id].append(row["Tên Chiều/Chỉ tiêu"])
                    rpt_mea_dict[rpt_id].append(row["Chỉ tiêu cơ sở"])

                elif row["Loại"] == "Chỉ tiêu phái sinh" and pd.notna(row["Chiều cơ sở"]):
                    data_dim.append({
                        'DIM name': row['Chiều cơ sở'],
                        'Thuộc tính': row['Chi tiết'],
                    })
                    rpt_mea_dict[rpt_id].append(row["Tên Chiều/Chỉ tiêu"])
                    rpt_dim_dict[rpt_id].append(row["Chiều cơ sở"])

    list_dim = pd.concat([list_dim, pd.DataFrame(data_dim)], ignore_index=True).drop_duplicates()
    list_mea = pd.concat([list_mea, pd.DataFrame(data_mea)], ignore_index=True).drop_duplicates()

    list_dim.sort_values(by='DIM name', inplace=True)
    list_mea.sort_values(by='MEA name', inplace=True)

    # Lọc ra các giá trị bị trùng lặp
    dim_values = sorted(set([item for sublist in rpt_dim_dict.values() for item in sublist]))
    mea_values = sorted(set([item for sublist in rpt_mea_dict.values() for item in sublist]))

    # Tạo một DataFrame với các sheet là hàng và các giá trị là cột
    rpt_dim_df = pd.DataFrame(index=rpt_dim_dict.keys(), columns=dim_values)
    rpt_mea_df = pd.DataFrame(index=rpt_mea_dict.keys(), columns=mea_values)

    # Tạo một DataFrame rỗng cho ma trận quan hệ giữa dim và mea
    dim_values = rpt_dim_df.columns
    mea_values = rpt_mea_df.columns
    dim_mea_df = pd.DataFrame(index=dim_values, columns=mea_values)

    # Đánh dấu "x" nếu sheet có chứa giá trị
    for report, values in rpt_dim_dict.items():
        for value in values:
            rpt_dim_df.loc[report, value] = 'x'

    for report, values in rpt_mea_dict.items():
        for value in values:
            rpt_mea_df.loc[report, value] = 'x'

    # Điền vào ma trận dim-mea dựa trên quan hệ report-dim và report-mea
    for report in rpt_dim_df.index:
        dims_in_report = rpt_dim_df.loc[report][rpt_dim_df.loc[report] == 'x'].index
        meas_in_report = rpt_mea_df.loc[report][rpt_mea_df.loc[report] == 'x'].index
        
        for dim in dims_in_report:
            for mea in meas_in_report:
                dim_mea_df.loc[dim, mea] = 'x'

    # Điền NaN thành chuỗi rỗng (nếu cần)
    dim_mea_df.fillna('', inplace=True)
    rpt_dim_df.fillna('', inplace=True)
    rpt_mea_df.fillna('', inplace=True)

    # Mở file Excel và kiểm tra các sheet có tồn tại không
    workbook = load_workbook(file_path)
    existing_sheets = workbook.sheetnames

    # Ghi mới các sheet với version tiếp theo
    sheet_names = ['List DIM', 'List MEA', 'Matrix RPT_DIM', 'Matrix RPT_MEA', 'Matrix MEA_DIM']
    versioned_sheet_names = []
    for sheet_name in sheet_names:
        new_name = get_next_version(sheet_name, existing_sheets)
        versioned_sheet_names.append(new_name)
        existing_sheets.append(new_name)

    # Lưu lại workbook
    # workbook.save(file_path)

    # Ghi các DataFrame vào các sheet riêng
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
        list_dim.to_excel(writer, sheet_name=versioned_sheet_names[0], index=False)
        list_mea.to_excel(writer, sheet_name=versioned_sheet_names[1], index=False)
        rpt_dim_df.to_excel(writer, sheet_name=versioned_sheet_names[2], index=True)
        rpt_mea_df.to_excel(writer, sheet_name=versioned_sheet_names[3], index=True)
        dim_mea_df.to_excel(writer, sheet_name=versioned_sheet_names[4], index=True)
        
    versions = versioned_sheet_names[0].split('_v')[-1] 

    print("\nĐã xuất thành công các sheet ma trận chỉ tiêu vào file Excel.")
    print(f"Phiên bản hiện tại: version {versions[0]}.\n")